"""
Keyword Landscape Analysis Script
Processes venue-keyword-research.xlsx and Keyword Research 01.xlsx
Generates structured data for the analysis report
"""
import openpyxl
import json
import re
from collections import defaultdict

# ──────────────────────────────────────────────
# 1. LOAD ALL DATA
# ──────────────────────────────────────────────

def safe_float(val):
    """Convert a value to float, handling strings like '1.0K', '22K' etc."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '')
    if s in ('N/A', '?', '', '0–10', '0\u201310'):
        return 0
    # Handle K suffix
    if s.upper().endswith('K'):
        try:
            return float(s[:-1]) * 1000
        except:
            return 0
    try:
        return float(s)
    except:
        return 0

def load_venue_keyword_research():
    """Load venue-keyword-research.xlsx"""
    wb = openpyxl.load_workbook('.strategy/venue-keyword-research.xlsx', read_only=True, data_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    results = []
    for r in rows[1:]:
        if r[0]:
            results.append({
                'keyword': str(r[0]).strip().lower(),
                'sv': safe_float(r[1]),
                'kd': safe_float(r[2]),
                'category_url': str(r[3]) if r[3] else '',
                'source': 'venue-keyword-research'
            })
    return results

def load_keyword_research_01():
    """Load all sheets from Keyword Research 01.xlsx"""
    wb = openpyxl.load_workbook('.strategy/Keyword Research 01.xlsx', read_only=True, data_only=True)
    all_data = {}

    # Wedding Venue Types
    ws = wb['Wedding Venue Types ']
    rows = list(ws.iter_rows(values_only=True))
    venue_types = []
    for r in rows[1:]:
        if r[0] and str(r[0]).strip():
            venue_types.append({
                'keyword': str(r[0]).strip().lower(),
                'kd': safe_float(r[1]),
                'sv_us': safe_float(r[2]),
                'gsv': safe_float(r[6]) if len(r) > 6 else 0,
                'tp': safe_float(r[7]) if len(r) > 7 else 0,
                'source': 'venue_types'
            })
    all_data['venue_types'] = venue_types

    # Wedding Venue Styles
    ws = wb['Wedding Venue styles ']
    rows = list(ws.iter_rows(values_only=True))
    venue_styles = []
    for r in rows[1:]:
        if r[0] and str(r[0]).strip():
            venue_styles.append({
                'keyword': str(r[0]).strip().lower(),
                'kd': safe_float(r[1]),
                'sv_us': safe_float(r[2]),
                'gsv': safe_float(r[6]) if len(r) > 6 else 0,
                'tp': safe_float(r[7]) if len(r) > 7 else 0,
                'source': 'venue_styles'
            })
    all_data['venue_styles'] = venue_styles

    # France Regions Keywords (manual)
    ws = wb['France Regions Keywords (manual']
    rows = list(ws.iter_rows(values_only=True))
    regions_manual = []
    for r in rows[1:]:
        if r[0] and str(r[0]).strip():
            regions_manual.append({
                'keyword': str(r[0]).strip().lower(),
                'region': str(r[1]).strip() if len(r) > 1 and r[1] else '',
                'city': str(r[2]).strip() if len(r) > 2 and r[2] else '',
                'type': str(r[3]).strip() if len(r) > 3 and r[3] else '',
                'sv_us': safe_float(r[4]) if len(r) > 4 else 0,
                'gsv': safe_float(r[5]) if len(r) > 5 else 0,
                'tp': safe_float(r[6]) if len(r) > 6 else 0,
                'source': 'regions_manual'
            })
    all_data['regions_manual'] = regions_manual

    # AreasDepartments
    ws = wb['AreasDepartments']
    rows = list(ws.iter_rows(values_only=True))
    areas = []
    for r in rows[1:]:
        if r[0] and str(r[0]).strip():
            areas.append({
                'keyword': str(r[0]).strip().lower(),
                'kd': safe_float(r[1]),
                'sv_us': safe_float(r[2]),
                'gsv': safe_float(r[3]) if len(r) > 3 else 0,
                'tp': safe_float(r[4]) if len(r) > 4 else 0,
                'gtp': safe_float(r[5]) if len(r) > 5 else 0,
                'source': 'areas_departments'
            })
    all_data['areas_departments'] = areas

    # Latest to be merged (regions)
    ws = wb['Latest to be merged (regions) ']
    rows = list(ws.iter_rows(values_only=True))
    merged = []
    for r in rows[2:]:
        if len(r) > 3 and r[0] and str(r[0]).strip() and str(r[0]).strip() != 'Keyword':
            merged.append({
                'keyword': str(r[0]).strip().lower(),
                'intents': str(r[1]) if len(r) > 1 and r[1] else '',
                'kd': safe_float(r[2]),
                'sv_us': safe_float(r[3]),
                'gsv': safe_float(r[7]) if len(r) > 7 else 0,
                'tp': safe_float(r[8]) if len(r) > 8 else 0,
                'gtp': safe_float(r[9]) if len(r) > 9 else 0,
                'source': 'latest_merged'
            })
    all_data['latest_merged'] = merged

    # Chateau - fresh data
    ws = wb['Chateau - fresh data']
    rows = list(ws.iter_rows(values_only=True))
    chateau_fresh = []
    for r in rows[1:]:
        if r[0] and str(r[0]).strip():
            chateau_fresh.append({
                'keyword': str(r[0]).strip().lower(),
                'sv_us': safe_float(r[1]),
                'gsv': safe_float(r[2]),
                'wedding_variant': str(r[3]).strip() if len(r) > 3 and r[3] else '',
                'priority': str(r[4]).strip() if len(r) > 4 and r[4] else '',
                'kd': safe_float(r[6]) if len(r) > 6 else 0,
                'wedding_volume': safe_float(r[8]) if len(r) > 8 else 0,
                'source': 'chateau_fresh'
            })
    all_data['chateau_fresh'] = chateau_fresh

    # Chateau - old filtered
    ws = wb['Chateau - old filtered ']
    rows = list(ws.iter_rows(values_only=True))
    chateau_old = []
    for r in rows[2:]:  # skip title + header
        if len(r) > 1 and r[0] and str(r[0]).strip():
            chateau_old.append({
                'keyword': str(r[0]).strip().lower(),
                'sv': safe_float(r[1]),
                'is_french': str(r[2]).strip().lower() == 'y' if len(r) > 2 and r[2] else False,
                'category': str(r[3]).strip() if len(r) > 3 and r[3] else '',
                'kd': safe_float(r[4]) if len(r) > 4 else 0,
                'source': 'chateau_old'
            })
    all_data['chateau_old'] = chateau_old

    # Sheet5 - City keywords
    ws = wb['Sheet5']
    rows = list(ws.iter_rows(values_only=True))
    city_kws = []
    for r in rows[1:]:
        if len(r) > 0 and r[0] and str(r[0]).strip() and str(r[0]).strip() != 'Keyword':
            city_kws.append({
                'keyword': str(r[0]).strip().lower(),
                'kd': safe_float(r[1]),
                'sv_us': safe_float(r[2]),
                'gsv': safe_float(r[3]) if len(r) > 3 else 0,
                'source': 'city_keywords'
            })
    all_data['city_keywords'] = city_kws

    # Sheet8 - Chateau suitability
    ws = wb['Sheet8']
    rows = list(ws.iter_rows(values_only=True))
    suitability = []
    for r in rows[1:]:
        if len(r) > 0 and r[0]:
            suitability.append({
                'name': str(r[0]).strip(),
                'suitable': str(r[1]).strip() if len(r) > 1 and r[1] else '',
                'location': str(r[2]).strip() if len(r) > 2 and r[2] else '',
                'region': str(r[3]).strip() if len(r) > 3 and r[3] else '',
            })
    all_data['chateau_suitability'] = suitability

    wb.close()
    return all_data


# ──────────────────────────────────────────────
# 2. CLASSIFICATION FUNCTIONS
# ──────────────────────────────────────────────

FRENCH_INDICATORS = [
    'mariage', 'salle', 'domaine', 'lieu', 'château', 'réception',
    'prix', 'location', 'île', 'fête', 'rêve', 'autour de moi',
    'hôtel', 'vidéo', 'cérémonie', 'traiteur', 'photographe',
    'fleuriste', 'décoration', 'musique', 'chauffeur', 'coiffeur',
    'maquillage', 'gîte', 'gite', 'ferme', 'vendre', 'vente',
    'louer', 'fortune', 'classement', 'nagui', 'brunch', 'gastronomique',
    'étoiles', 'etoiles', 'médiéval', 'dormir', 'visiter',
    'ile de france', 'bourgogne', 'franche-comté', 'rhône-alpes',
    'hauts-de-france', 'de mariage', 'pour mariage'
]

WEDDING_KEYWORDS = [
    'wedding', 'bride', 'bridal', 'groom', 'elopement', 'elope',
    'vow renewal', 'ceremony', 'reception', 'nuptial', 'honeymoon',
    'engagement', 'propose', 'proposal', 'marry', 'married',
    'getting married', 'destination wedding', 'intimate wedding'
]

VENUE_TYPE_MAP = {
    'château': ['chateau', 'château', 'chateaux', 'châteaux', 'castle'],
    'domaine': ['domaine', 'domain'],
    'villa': ['villa'],
    'bastide': ['bastide'],
    'mas': ['mas '],
    'manoir': ['manoir', 'manor'],
    'farmhouse': ['farmhouse', 'ferme', 'grange', 'farm'],
    'barn': ['barn'],
    'vineyard': ['vineyard', 'winery', 'wine estate', 'vignoble'],
    'hotel': ['hotel', 'hôtel', 'boutique hotel'],
    'garden': ['garden', 'orangerie', 'jardin'],
    'beach': ['beach', 'coastal', 'seaside', 'sea view'],
    'restaurant': ['restaurant'],
    'boat': ['boat', 'yacht', 'barge'],
}

REGION_MAP = {
    'Provence': ['provence', 'aix-en-provence', 'aix en provence', 'luberon', 'alpilles', 'vaucluse', 'var '],
    'French Riviera': ['riviera', 'côte d\'azur', 'cote d\'azur', 'nice ', 'cannes', 'st tropez', 'saint tropez', 'antibes', 'monaco', 'grasse', 'èze', 'eze'],
    'Île-de-France / Paris': ['paris', 'ile-de-france', 'île-de-france', 'versailles', 'fontainebleau', 'chantilly'],
    'Nouvelle-Aquitaine': ['bordeaux', 'dordogne', 'charente', 'cognac', 'biarritz', 'nouvelle-aquitaine', 'aquitaine', 'périgord', 'perigord', 'lot-et-garonne', 'gironde'],
    'Occitanie': ['occitanie', 'occitania', 'languedoc', 'toulouse', 'montpellier', 'carcassonne', 'nîmes', 'nimes', 'hérault'],
    'Loire Valley': ['loire', 'val de loire', 'touraine', 'anjou', 'amboise', 'chenonceau', 'chambord'],
    'Burgundy': ['burgundy', 'bourgogne', 'beaune', 'dijon', 'chablis'],
    'Normandy': ['normandy', 'normandie', 'deauville', 'honfleur', 'mont saint michel'],
    'Grand Est': ['alsace', 'champagne', 'lorraine', 'grand est', 'strasbourg', 'reims'],
    'Corsica': ['corsica', 'corse'],
    'Auvergne-Rhône-Alpes': ['auvergne', 'rhône-alpes', 'rhone-alpes', 'alps', 'lyon', 'annecy', 'chamonix'],
    'Brittany': ['brittany', 'bretagne'],
    'Hauts-de-France': ['hauts-de-france', 'picardy', 'picardie'],
    'South of France': ['south of france', 'south france', 'midi'],
    'South-West France': ['south-west france', 'southwest france', 'south west france'],
    'North of France': ['north of france', 'north france', 'northern france'],
}

FEATURE_MAP = {
    'pool': ['pool', 'piscine', 'swimming'],
    'accommodation': ['accommodation', 'sleep', 'stay', 'hébergement'],
    'all-inclusive': ['all-inclusive', 'all inclusive', 'package'],
    'exclusive-use': ['exclusive use', 'exclusive-use', 'private hire'],
    'external-caterers': ['external caterer', 'outside caterer', 'dry hire', 'dry-hire', 'bring your own caterer'],
    'no-curfew': ['no curfew', 'no-curfew', 'late night', 'party'],
    'pet-friendly': ['pet friendly', 'pet-friendly', 'dog friendly', 'dog-friendly'],
    'chapel': ['chapel', 'church', 'église'],
    'lgbtq': ['lgbtq', 'gay wedding', 'same sex', 'same-sex'],
    'eco': ['eco', 'sustainable', 'green wedding'],
}

CAPACITY_MAP = {
    'elopement': ['elopement', 'elope', 'just the two', 'for two'],
    'intimate': ['intimate', 'small wedding', 'micro wedding', 'tiny wedding'],
    'large': ['large wedding', 'big wedding', '200 guest', '300 guest'],
}

BUDGET_MAP = {
    'affordable': ['affordable', 'budget', 'cheap', 'under', 'low cost'],
    'luxury': ['luxury', 'luxurious', 'premium', 'exclusive'],
}

SEASONAL_MAP = {
    'summer': ['summer'],
    'winter': ['winter', 'christmas', 'new year'],
    'autumn': ['autumn', 'fall', 'harvest'],
    'spring': ['spring'],
}


def is_french(kw):
    kw_lower = kw.lower()
    for indicator in FRENCH_INDICATORS:
        if indicator in kw_lower:
            return True
    return False

def is_wedding_relevant(kw):
    kw_lower = kw.lower()
    for w in WEDDING_KEYWORDS:
        if w in kw_lower:
            return True
    # Also count venue-type + region combinations as wedding-relevant
    has_venue = any(any(v in kw_lower for v in variants) for variants in VENUE_TYPE_MAP.values())
    has_region = any(any(r in kw_lower for r in indicators) for indicators in REGION_MAP.values())
    if has_venue and has_region:
        return True
    return False

def classify_venue_type(kw):
    kw_lower = kw.lower()
    types = []
    for vtype, indicators in VENUE_TYPE_MAP.items():
        for ind in indicators:
            if ind in kw_lower:
                types.append(vtype)
                break
    return types

def classify_region(kw):
    kw_lower = kw.lower()
    regions = []
    for region, indicators in REGION_MAP.items():
        for ind in indicators:
            if ind in kw_lower:
                regions.append(region)
                break
    return regions

def classify_feature(kw):
    kw_lower = kw.lower()
    features = []
    for feat, indicators in FEATURE_MAP.items():
        for ind in indicators:
            if ind in kw_lower:
                features.append(feat)
                break
    return features

def classify_capacity(kw):
    kw_lower = kw.lower()
    for cap, indicators in CAPACITY_MAP.items():
        for ind in indicators:
            if ind in kw_lower:
                return cap
    return None

def classify_budget(kw):
    kw_lower = kw.lower()
    for bud, indicators in BUDGET_MAP.items():
        for ind in indicators:
            if ind in kw_lower:
                return bud
    return None

def classify_season(kw):
    kw_lower = kw.lower()
    for season, indicators in SEASONAL_MAP.items():
        for ind in indicators:
            if ind in kw_lower:
                return season
    return None


# ──────────────────────────────────────────────
# 3. MAIN ANALYSIS
# ──────────────────────────────────────────────

def main():
    print("Loading data...")
    file1_data = load_venue_keyword_research()
    file2_data = load_keyword_research_01()

    # Build master deduplicated keyword map (keyword -> best data)
    master = {}

    # Process File 1
    for item in file1_data:
        kw = item['keyword']
        if kw not in master or item['sv'] > master[kw].get('sv', 0):
            master[kw] = {
                'keyword': kw,
                'sv': item['sv'],
                'kd': item['kd'],
                'sources': [item['source']],
                'category_url': item['category_url'],
            }
        else:
            if item['source'] not in master[kw].get('sources', []):
                master[kw].setdefault('sources', []).append(item['source'])

    # Process File 2 - all sheets
    for sheet_name, items in file2_data.items():
        if sheet_name == 'chateau_suitability':
            continue  # Not keyword data
        for item in items:
            kw = item['keyword']
            sv = item.get('sv_us', item.get('sv', 0))
            gsv = item.get('gsv', 0)
            best_sv = max(sv, gsv)
            if kw not in master:
                master[kw] = {
                    'keyword': kw,
                    'sv': best_sv,
                    'sv_us': sv,
                    'gsv': gsv,
                    'kd': item.get('kd', 0),
                    'sources': [item['source']],
                }
            else:
                if item['source'] not in master[kw].get('sources', []):
                    master[kw].setdefault('sources', []).append(item['source'])
                # Update with better data
                if best_sv > master[kw].get('sv', 0):
                    master[kw]['sv'] = best_sv
                if gsv > master[kw].get('gsv', 0):
                    master[kw]['gsv'] = gsv
                if sv > master[kw].get('sv_us', 0):
                    master[kw]['sv_us'] = sv

    # Classify all keywords
    print(f"Classifying {len(master)} unique keywords...")
    for kw, data in master.items():
        data['is_french'] = is_french(kw)
        data['is_wedding'] = is_wedding_relevant(kw)
        data['venue_types'] = classify_venue_type(kw)
        data['regions'] = classify_region(kw)
        data['features'] = classify_feature(kw)
        data['capacity'] = classify_capacity(kw)
        data['budget'] = classify_budget(kw)
        data['season'] = classify_season(kw)

    # ── ANALYSIS OUTPUTS ──

    all_kws = list(master.values())
    total_kws = len(all_kws)
    english_kws = [k for k in all_kws if not k['is_french']]
    french_kws = [k for k in all_kws if k['is_french']]
    wedding_kws = [k for k in all_kws if k['is_wedding']]
    wedding_english = [k for k in wedding_kws if not k['is_french']]

    total_sv = sum(k['sv'] for k in all_kws)
    english_sv = sum(k['sv'] for k in english_kws)
    french_sv = sum(k['sv'] for k in french_kws)
    wedding_sv = sum(k['sv'] for k in wedding_kws)
    wedding_english_sv = sum(k['sv'] for k in wedding_english)

    print("\n" + "="*60)
    print("KEYWORD UNIVERSE OVERVIEW")
    print("="*60)
    print(f"Total unique keywords: {total_kws:,}")
    print(f"  English: {len(english_kws):,} ({len(english_kws)/total_kws*100:.0f}%)")
    print(f"  French: {len(french_kws):,} ({len(french_kws)/total_kws*100:.0f}%)")
    print(f"  Wedding-relevant: {len(wedding_kws):,} ({len(wedding_kws)/total_kws*100:.0f}%)")
    print(f"  Wedding + English: {len(wedding_english):,}")
    print(f"\nTotal SV: {total_sv:,.0f}")
    print(f"  English SV: {english_sv:,.0f} ({english_sv/total_sv*100:.0f}%)")
    print(f"  French SV: {french_sv:,.0f} ({french_sv/total_sv*100:.0f}%)")
    print(f"  Wedding-relevant SV: {wedding_sv:,.0f}")
    print(f"  Wedding + English SV: {wedding_english_sv:,.0f}")

    # Volume distribution
    tiers = {'10K+': 0, '1K-10K': 0, '100-999': 0, '10-99': 0, '<10': 0}
    for k in wedding_english:
        sv = k['sv']
        if sv >= 10000: tiers['10K+'] += 1
        elif sv >= 1000: tiers['1K-10K'] += 1
        elif sv >= 100: tiers['100-999'] += 1
        elif sv >= 10: tiers['10-99'] += 1
        else: tiers['<10'] += 1
    print(f"\nWedding English SV distribution:")
    for tier, count in tiers.items():
        print(f"  {tier}: {count} keywords")

    # ── VENUE TYPE ANALYSIS ──
    print("\n" + "="*60)
    print("VENUE TYPE KEYWORD DEMAND")
    print("="*60)

    # Combine data from venue_types sheet + main keyword data
    type_clusters = {}
    for vtype in VENUE_TYPE_MAP:
        matching = [k for k in english_kws if vtype in k.get('venue_types', [])]
        wedding_matching = [k for k in matching if k['is_wedding']]
        type_clusters[vtype] = {
            'total_kws': len(matching),
            'wedding_kws': len(wedding_matching),
            'total_sv': sum(k['sv'] for k in matching),
            'wedding_sv': sum(k['sv'] for k in wedding_matching),
            'avg_kd': sum(k['kd'] for k in matching if k['kd'] > 0) / max(1, len([k for k in matching if k['kd'] > 0])),
            'top_kws': sorted(wedding_matching, key=lambda x: x['sv'], reverse=True)[:5]
        }

    # Also pull in the venue_types sheet data
    for item in file2_data.get('venue_types', []):
        print(f"  Type sheet: {item['keyword']:30s}  SV(US)={item['sv_us']:>8}  GSV={item['gsv']:>8}  KD={item['kd']:>4}")

    print("\nVenue Type Clusters (English, wedding-relevant):")
    for vtype, data in sorted(type_clusters.items(), key=lambda x: x[1]['wedding_sv'], reverse=True):
        if data['wedding_sv'] > 0:
            print(f"\n  {vtype.upper()} — {data['wedding_kws']} kws, SV={data['wedding_sv']:,.0f}, Avg KD={data['avg_kd']:.0f}")
            for k in data['top_kws'][:3]:
                print(f"    SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # ── REGION ANALYSIS ──
    print("\n" + "="*60)
    print("REGIONAL KEYWORD DEMAND")
    print("="*60)

    region_clusters = {}
    for region in REGION_MAP:
        matching = [k for k in english_kws if region in k.get('regions', [])]
        wedding_matching = [k for k in matching if k['is_wedding']]
        region_clusters[region] = {
            'total_kws': len(matching),
            'wedding_kws': len(wedding_matching),
            'total_sv': sum(k['sv'] for k in matching),
            'wedding_sv': sum(k['sv'] for k in wedding_matching),
            'avg_kd': sum(k['kd'] for k in matching if k['kd'] > 0) / max(1, len([k for k in matching if k['kd'] > 0])),
            'top_kws': sorted(wedding_matching, key=lambda x: x['sv'], reverse=True)[:5]
        }

    # Also add data from regions_manual and areas_departments
    print("\nRegion Clusters (all English keywords):")
    for region, data in sorted(region_clusters.items(), key=lambda x: x[1]['total_sv'], reverse=True):
        if data['total_sv'] > 0:
            print(f"\n  {region:30s} — Total: {data['total_kws']} kws / SV={data['total_sv']:>8,.0f}  |  Wedding: {data['wedding_kws']} kws / SV={data['wedding_sv']:>8,.0f}  |  Avg KD={data['avg_kd']:.0f}")
            for k in data['top_kws'][:3]:
                print(f"    SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # Add regions_manual data by region
    print("\n\nRegions Manual Sheet - Aggregated by Region:")
    region_manual_agg = defaultdict(lambda: {'kws': 0, 'sv_us': 0, 'gsv': 0})
    for item in file2_data.get('regions_manual', []):
        reg = item['region']
        if reg:
            region_manual_agg[reg]['kws'] += 1
            region_manual_agg[reg]['sv_us'] += item['sv_us']
            region_manual_agg[reg]['gsv'] += item['gsv']
    for reg, data in sorted(region_manual_agg.items(), key=lambda x: x[1]['gsv'], reverse=True):
        print(f"  {reg:40s}  KWs={data['kws']:>4}  SV(US)={data['sv_us']:>8,.0f}  GSV={data['gsv']:>10,.0f}")

    # ── CHÂTEAU DEEP DIVE ──
    print("\n" + "="*60)
    print("CHÂTEAU DEEP DIVE")
    print("="*60)

    chateau_fresh = file2_data.get('chateau_fresh', [])
    print(f"Individual château keywords: {len(chateau_fresh)}")

    # Priority breakdown
    priority_counts = defaultdict(int)
    priority_sv = defaultdict(float)
    for item in chateau_fresh:
        p = item['priority'] or 'Unknown'
        priority_counts[p] += 1
        priority_sv[p] += item['sv_us']
    print("\nPriority Distribution:")
    for p in ['High', 'Medium', 'Low', 'Unknown']:
        if priority_counts[p]:
            print(f"  {p}: {priority_counts[p]} châteaux, Total US SV={priority_sv[p]:,.0f}")

    # Top wedding-searched châteaux
    wedding_chateaux = sorted([c for c in chateau_fresh if c['wedding_variant']], key=lambda x: x['sv_us'], reverse=True)
    print(f"\nTop 20 Château Names by Search Volume (with wedding variant):")
    for c in wedding_chateaux[:20]:
        print(f"  US={c['sv_us']:>6,.0f}  Global={c['gsv']:>8,.0f}  KD={c['kd']:>3.0f}  {c['keyword']}  -> \"{c['wedding_variant']}\"")

    # Total château wedding cluster volume
    chateau_wedding_kws = [k for k in all_kws if 'château' in k.get('venue_types', []) and k['is_wedding']]
    print(f"\nTotal 'château + wedding' keyword cluster: {len(chateau_wedding_kws)} kws, SV={sum(k['sv'] for k in chateau_wedding_kws):,.0f}")

    # ── LONG-TAIL & NICHE ──
    print("\n" + "="*60)
    print("LONG-TAIL & NICHE OPPORTUNITIES")
    print("="*60)

    # Feature keywords
    print("\nFeature Keywords (English, wedding-relevant):")
    for feat in FEATURE_MAP:
        matching = [k for k in wedding_english if feat in k.get('features', [])]
        total_sv = sum(k['sv'] for k in matching)
        if matching:
            print(f"\n  {feat.upper()} — {len(matching)} kws, SV={total_sv:,.0f}")
            for k in sorted(matching, key=lambda x: x['sv'], reverse=True)[:3]:
                print(f"    SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # Capacity
    print("\nCapacity Keywords:")
    for cap in CAPACITY_MAP:
        matching = [k for k in wedding_english if k.get('capacity') == cap]
        total_sv = sum(k['sv'] for k in matching)
        if matching:
            print(f"\n  {cap.upper()} — {len(matching)} kws, SV={total_sv:,.0f}")
            for k in sorted(matching, key=lambda x: x['sv'], reverse=True)[:3]:
                print(f"    SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # Budget
    print("\nBudget Keywords:")
    for bud in BUDGET_MAP:
        matching = [k for k in wedding_english if k.get('budget') == bud]
        total_sv = sum(k['sv'] for k in matching)
        if matching:
            print(f"\n  {bud.upper()} — {len(matching)} kws, SV={total_sv:,.0f}")
            for k in sorted(matching, key=lambda x: x['sv'], reverse=True)[:3]:
                print(f"    SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # Seasonal
    print("\nSeasonal Keywords:")
    for season in SEASONAL_MAP:
        matching = [k for k in wedding_english if k.get('season') == season]
        total_sv = sum(k['sv'] for k in matching)
        if matching:
            print(f"\n  {season.upper()} — {len(matching)} kws, SV={total_sv:,.0f}")
            for k in sorted(matching, key=lambda x: x['sv'], reverse=True)[:3]:
                print(f"    SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # ── COMPETITIVE GAPS ──
    print("\n" + "="*60)
    print("COMPETITIVE KEYWORD INTELLIGENCE")
    print("="*60)

    # Analyze category URLs from file 1
    competitor_domains = defaultdict(lambda: {'kws': 0, 'sv': 0, 'top_kws': []})
    for item in file1_data:
        url = item.get('category_url', '')
        if url and url.startswith('http'):
            # Extract domain
            domain = url.split('/')[2] if len(url.split('/')) > 2 else url
            competitor_domains[domain]['kws'] += 1
            competitor_domains[domain]['sv'] += item['sv']
            competitor_domains[domain]['top_kws'].append((item['keyword'], item['sv']))

    print("Top Competitor Domains (by keyword count):")
    for domain, data in sorted(competitor_domains.items(), key=lambda x: x[1]['kws'], reverse=True)[:15]:
        top_3 = sorted(data['top_kws'], key=lambda x: x[1], reverse=True)[:3]
        print(f"\n  {domain}")
        print(f"    Keywords: {data['kws']}, Total SV: {data['sv']:,.0f}")
        for kw, sv in top_3:
            print(f"      SV={sv:>6,.0f}  {kw}")

    # ── FRENCH LANGUAGE INTELLIGENCE ──
    print("\n" + "="*60)
    print("FRENCH LANGUAGE KEYWORD INTELLIGENCE")
    print("="*60)

    top_french = sorted(french_kws, key=lambda x: x['sv'], reverse=True)[:30]
    print("Top 30 French Keywords:")
    for k in top_french:
        print(f"  SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")

    # French wedding keywords specifically
    french_wedding = [k for k in french_kws if any(w in k['keyword'] for w in ['mariage', 'noce', 'épouser'])]
    print(f"\nFrench 'mariage' keywords: {len(french_wedding)}, Total SV={sum(k['sv'] for k in french_wedding):,.0f}")
    for k in sorted(french_wedding, key=lambda x: x['sv'], reverse=True)[:15]:
        print(f"  SV={k['sv']:>6,.0f}  {k['keyword']}")

    # ── QUICK WINS: High SV + Low KD ──
    print("\n" + "="*60)
    print("QUICK WINS: HIGH SV + LOW KD (English, Wedding)")
    print("="*60)

    quick_wins = [k for k in wedding_english if k['sv'] >= 50 and k['kd'] > 0 and k['kd'] <= 15]
    quick_wins_sorted = sorted(quick_wins, key=lambda x: x['sv'], reverse=True)
    print(f"Keywords with SV >= 50 and KD <= 15:")
    for k in quick_wins_sorted[:30]:
        types = ', '.join(k['venue_types']) if k['venue_types'] else '-'
        regs = ', '.join(k['regions']) if k['regions'] else '-'
        print(f"  SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  Types=[{types}]  Regions=[{regs}]  {k['keyword']}")

    # ── TOP WEDDING KEYWORDS OVERALL ──
    print("\n" + "="*60)
    print("TOP 50 WEDDING KEYWORDS (English, by SV)")
    print("="*60)
    top_wedding = sorted(wedding_english, key=lambda x: x['sv'], reverse=True)
    for k in top_wedding[:50]:
        types = ', '.join(k['venue_types']) if k['venue_types'] else '-'
        regs = ', '.join(k['regions']) if k['regions'] else '-'
        feats = ', '.join(k['features']) if k['features'] else '-'
        print(f"  SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  Type=[{types}]  Region=[{regs}]  Feature=[{feats}]  {k['keyword']}")

    # ── HUB PAGE VALIDATION ──
    print("\n" + "="*60)
    print("HUB PAGE KEYWORD VALIDATION (P0 Pages)")
    print("="*60)

    p0_pages = [
        ("Château Wedding Venues in France", ["chateau wedding", "château wedding", "castle wedding france", "chateau wedding france", "chateau wedding venues in france", "château wedding venues in france"]),
        ("Wedding Venues in Provence", ["wedding venues in provence", "provence wedding venues", "wedding venue provence", "wedding provence"]),
        ("Wedding Venues in the South of France", ["south of france wedding venues", "wedding venues south of france", "wedding venues in the south of france", "south of france wedding"]),
        ("Château Wedding Venues in Provence", ["chateau wedding provence", "château wedding provence", "chateau venues provence", "chateau wedding venues in provence"]),
        ("Wedding Venues in France with a Pool", ["wedding venues france pool", "wedding venues with pool france", "pool wedding venue france"]),
        ("Vineyard Wedding Venues in France", ["vineyard wedding france", "vineyard wedding venues france", "winery wedding france", "vineyard wedding venues in france"]),
        ("Wedding Venues near Paris", ["wedding venues near paris", "paris wedding venues", "wedding venues paris", "wedding venue paris"]),
        ("Intimate Wedding Venues in France", ["intimate wedding france", "intimate wedding venues france", "small wedding france", "intimate wedding venues in france"]),
        ("Wedding Venues on the French Riviera", ["french riviera wedding venues", "wedding venues french riviera", "riviera wedding venues"]),
        ("All-Inclusive Wedding Venues in France", ["all-inclusive wedding france", "all inclusive wedding venues france"]),
    ]

    for page_title, target_kws in p0_pages:
        matching_kws = []
        for target in target_kws:
            if target in master:
                matching_kws.append(master[target])
        # Also fuzzy match
        for kw, data in master.items():
            if any(t in kw for t in target_kws[:2]) and data not in matching_kws:
                matching_kws.append(data)

        total_sv = sum(k['sv'] for k in matching_kws)
        avg_kd = sum(k['kd'] for k in matching_kws if k['kd'] > 0) / max(1, len([k for k in matching_kws if k['kd'] > 0]))
        print(f"\n  {page_title}")
        print(f"    Matching keywords: {len(matching_kws)}, Combined SV: {total_sv:,.0f}, Avg KD: {avg_kd:.0f}")
        for k in sorted(matching_kws, key=lambda x: x['sv'], reverse=True)[:5]:
            print(f"      SV={k['sv']:>6,.0f}  KD={k['kd']:>3.0f}  {k['keyword']}")


if __name__ == '__main__':
    main()
